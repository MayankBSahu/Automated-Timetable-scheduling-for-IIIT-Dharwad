import pandas as pd
import json
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Set fixed seed for consistent but unique randomization
random.seed(42)

# Load data
with open("data/time_slots.json") as f:
    slots_data = json.load(f)["time_slots"]

slot_keys = [f"{slot['start'].strip()}-{slot['end'].strip()}" for slot in slots_data]

# ---------------------------------------
# Helper functions and utilities
# ---------------------------------------

def calc_slot_duration(slot):
    start, end = slot.split("-")
    h1, m1 = map(int, start.split(":"))
    h2, m2 = map(int, end.split(":"))
    return (h2 + m2 / 60) - (h1 + m1 / 60)

time_durations = {s: calc_slot_duration(s) for s in slot_keys}

# Generate random pastel colors dynamically
def random_pastel():
    return "%02X%02X%02X" % (
        random.randint(180, 255),
        random.randint(180, 255),
        random.randint(180, 255)
    )

color_palette = [random_pastel() for _ in range(20)]

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# ---------------------------------------
# Load course and room data
# ---------------------------------------
courses_df = pd.read_csv("data/courses.csv").to_dict(orient="records")
rooms_df = pd.read_csv("data/rooms.csv")  # must have columns: Room_ID, Type
classrooms = rooms_df[rooms_df["Type"].str.lower() == "classroom"]["Room_ID"].tolist()
labs = rooms_df[rooms_df["Type"].str.lower() == "lab"]["Room_ID"].tolist()

week_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded_slots = ["07:30-09:00", "13:15-14:00", "17:30-18:30"]
MAX_ATTEMPTS = 10

# ---------------------------------------
# Core allocation functions
# ---------------------------------------

def find_free_blocks(tt_frame, day):
    free_groups = []
    temp_group = []
    for slot in slot_keys:
        if tt_frame.at[day, slot] == "" and slot not in excluded_slots:
            temp_group.append(slot)
        else:
            if temp_group:
                free_groups.append(temp_group)
                temp_group = []
    if temp_group:
        free_groups.append(temp_group)
    return free_groups


room_allocations = {}

def place_session(tt_frame, faculty_schedule, day, faculty, course_code, duration_hours, session_type="L", is_elective=False):
    free_groups = find_free_blocks(tt_frame, day)
    for group in free_groups:
        total = sum(time_durations[s] for s in group)
        if total >= duration_hours:
            slots_to_fill = []
            duration_accum = 0
            for s in group:
                slots_to_fill.append(s)
                duration_accum += time_durations[s]
                if duration_accum >= duration_hours:
                    break

            if not is_elective:
                if course_code in room_allocations:
                    room = room_allocations[course_code]
                else:
                    if session_type == "P":
                        if not labs:
                            print(f"No labs available for {course_code}")
                            return False
                        room = random.choice(labs)
                    else:
                        if not classrooms:
                            print(f"No classrooms available for {course_code}")
                            return False
                        room = random.choice(classrooms)
                    room_allocations[course_code] = room

            for s in slots_to_fill:
                if session_type == "L":
                    tt_frame.at[day, s] = f"{course_code} ({room})" if not is_elective else course_code
                elif session_type == "T":
                    tt_frame.at[day, s] = f"{course_code}T ({room})" if not is_elective else f"{course_code}T"
                elif session_type == "P":
                    tt_frame.at[day, s] = f"{course_code} (Lab-{room})" if not is_elective else course_code

            if faculty:
                faculty_schedule[day].append(faculty)
            return True
    return False


def style_and_merge_cells(filename):
    wb = load_workbook(filename)
    ws = wb.active

    course_colors = {}
    available_colors = color_palette.copy()
    random.shuffle(available_colors)

    for row_idx in range(2, ws.max_row + 1):
        col_start = 2
        while col_start <= ws.max_column:
            value = ws.cell(row=row_idx, column=col_start).value
            if not value:
                col_start += 1
                continue

            merge_cols = [col_start]
            if "(" in str(value):
                if "Lab" in value:
                    session_dur = 2
                elif "T" in value:
                    session_dur = 1
                else:
                    session_dur = 1.5
            else:
                session_dur = 1.5

            total_dur = time_durations[slot_keys[col_start - 2]]
            next_col = col_start + 1
            while next_col <= ws.max_column:
                next_value = ws.cell(row=row_idx, column=next_col).value
                if next_value == value:
                    total_dur += time_durations[slot_keys[next_col - 2]]
                    merge_cols.append(next_col)
                    if total_dur >= session_dur:
                        break
                    next_col += 1
                else:
                    break

            if len(merge_cols) > 1:
                ws.merge_cells(start_row=row_idx, start_column=merge_cols[0],
                               end_row=row_idx, end_column=merge_cols[-1])

            cell = ws.cell(row=row_idx, column=merge_cols[0])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

            raw_name = str(value).split()[0].replace("T", "")
            if raw_name not in course_colors:
                if available_colors:
                    course_colors[raw_name] = available_colors.pop()
                else:
                    r = lambda: random.randint(150, 255)
                    course_colors[raw_name] = f"{r():02X}{r():02X}{r():02X}"

            fill_color = course_colors[raw_name]
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for col in merge_cols:
                ws.cell(row=row_idx, column=col).border = thin_border

            col_start = merge_cols[-1] + 1

    wb.save(filename)


def create_timetable(course_list, filename):
    tt_frame = pd.DataFrame("", index=week_days, columns=slot_keys)
    faculty_schedule = {day: [] for day in week_days}
    global room_allocations
    room_allocations = {}

    electives = [c for c in course_list if str(c.get("Elective", 0)) == "1"]
    mandatory = [c for c in course_list if str(c.get("Elective", 0)) != "1"]

    if electives:
        chosen_elective = random.choice(electives)
        elective_course = {
            "Course_Code": "Elective",
            "Faculty": chosen_elective.get("Faculty", ""),
            "L-T-P-S-C": chosen_elective["L-T-P-S-C"]
        }
        mandatory.append(elective_course)

    for course in mandatory:
        faculty = str(course.get("Faculty", "")).strip()
        course_code = str(course["Course_Code"]).strip()
        is_elective = True if course_code == "Elective" else False

        try:
            L, T, P, S, C = map(int, [x.strip() for x in course["L-T-P-S-C"].split("-")])
        except:
            continue

        lecture_hours_remaining = L
        attempts = 0
        while lecture_hours_remaining > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in week_days:
                if lecture_hours_remaining <= 0 or (faculty and faculty in faculty_schedule[day]):
                    continue
                alloc_hours = min(1.5, lecture_hours_remaining)
                if place_session(tt_frame, faculty_schedule, day, faculty, course_code, alloc_hours, "L", is_elective):
                    lecture_hours_remaining -= alloc_hours
                    break

        tutorial_hours_remaining = T
        attempts = 0
        while tutorial_hours_remaining > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in week_days:
                if tutorial_hours_remaining <= 0 or (faculty and faculty in faculty_schedule[day]):
                    continue
                if place_session(tt_frame, faculty_schedule, day, faculty, course_code, 1, "T", is_elective):
                    tutorial_hours_remaining -= 1
                    break
                    
        practical_hours_remaining = P
        attempts = 0
        while practical_hours_remaining > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in week_days:
                if practical_hours_remaining <= 0 or (faculty and faculty in faculty_schedule[day]):
                    continue
                alloc_hours = min(2, practical_hours_remaining) if practical_hours_remaining >= 2 else practical_hours_remaining
                if place_session(tt_frame, faculty_schedule, day, faculty, course_code, alloc_hours, "P", is_elective):
                    practical_hours_remaining -= alloc_hours
                    break

    for day in week_days:
        for slot in excluded_slots:
            if slot in tt_frame.columns:
                tt_frame.at[day, slot] = ""

    tt_frame.to_excel(filename, index=True)
    style_and_merge_cells(filename)
    print(f"Saved styled timetable to {filename}")

# ---------------------------------------
# Generate first and second half timetables
# ---------------------------------------
courses_first_half = [c for c in courses_df if str(c.get("Semester_Half")).strip() in ["1", "0"]]
courses_second_half = [c for c in courses_df if str(c.get("Semester_Half")).strip() in ["2", "0"]]

create_timetable(courses_first_half, "timetable_first_half.xlsx")
create_timetable(courses_second_half, "timetable_second_half.xlsx")
